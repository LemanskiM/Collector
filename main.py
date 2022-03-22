from openpyxl import load_workbook
import os
import csv
import pandas as pd

# path adress
# import folders name
path_folders = "C:\\Users\\LG_ML\\Desktop\Pobieracz\\form\\"
# save path adress
path_to_save = "C:\\Users\\LG_ML\\Desktop\Pobieracz\\"

# list to append files and folder name to create path
list_of_files = []
list_of_folders = []

# lists to merge and filtr .xlsm
ALL_PATH = []
ALL_PATH_EXCEL = []

def takeFolderList(path):
    """
    take folder list from folders
    path - path_folders adress
    """
    l = list(os.listdir(path))
    list_of_folders.append(l)

def sort():
    """
    taking files in folders
    """
    for i in list_of_folders[0]:
        l = list(os.listdir(path_folders+i))
        list_of_files.append(l)

def listOfAllAdress():
    """
    creating list
    """
    list_of_folders1 = list_of_folders[0]
    list = 0
    for i in list_of_files:
        for test in i:
            ALL_PATH.append(list_of_folders1[list]+"\\"+test)
        list+=1

def onlyExcel():
    """
    def is picking only excel files
    """
    for i in ALL_PATH:
        if i[-4:] == "xlsx":
            ALL_PATH_EXCEL.append(i)


"""
# Pop from file with openpyxl lib

VALUES_WN = []
VALUES_PD = []
def wnPop():
    for path in ALL_PATH_EXCEL:
        if path[:5] == "1.2.1":
            path = load_workbook(path_folders + path)
            sheet = path["RAPORT_PL_EN"]
            for row in sheet.iter_rows(min_row=22, min_col=2, max_row=41, max_col=6):
                for cell in row:
                     VALUES_WN.append(cell.value)
"""

def exportCsv():
    """
    is exporting data to new csv files
    """
    for path in ALL_PATH_EXCEL:
        if path[:5] == "1.2.1":
            df = pd.read_excel(path_folders+path, sheet_name = "RAPORT_PL_EN").iloc[20:39,1:7]
            df["adress"] = path
            df.to_csv(path_to_save + "wilgotnosc_naturalna_zebrane.csv", mode='a', header=False)
        if path[:5] == "1.3.1":
            df = pd.read_excel(path_folders+path, sheet_name = "RAPORT_PL_EN").iloc[20:39,1:7]
            df["adress"] = path
            df.to_csv(path_to_save + "gęstosc_objetosciowa_zebrane.csv", mode='a', header=False)
        if path[:5] == "1.5.1":
            df = pd.read_excel(path_folders+path, sheet_name = "RAPORT_PL_EN").iloc[20:39,1:7]
            df["adress"] = path
            df.to_csv(path_to_save + "zawartosc_substancji_organicznej_zebrane.csv", mode='a', header=False)
        if path[:5] == "1.3.3":
            df = pd.read_excel(path_folders+path, sheet_name = "RAPORT").iloc[20:42,1:6]
            df["adress"] = path
            df.to_csv(path_to_save + "gęstość_właściwa_zebrane.csv", mode='a', header=False)
        if path[:5] == "1.5.2":
            df = pd.read_excel(path_folders+path, sheet_name = "RAPORT_PL_EN").iloc[20:51,1:7]
            df["adress"] = path
            df.to_csv(path_to_save + "badanie_przewodnosci_termicznej_zebrane.csv", mode='a', header=False)
        if path[:5] == "1.4.1":
            df = pd.read_excel(path_folders+path, sheet_name = "RAPORT_PL_EN").iloc[19:20,1:10]
            df["adress"] = path
            df.to_csv(path_to_save + "Sianie_zebrane.csv", mode='a', header=False)

"""
# testing sieve

for path in ALL_PATH_EXCEL:
    if path[:5] == "1.4.1":
        df = pd.read_excel(path_folders+path, sheet_name = "RAPORT_PL_EN")
        print(df)
#        df["adress"] = path
#        df.to_csv(path_to_save + "Sianie_zebrane.csv", mode='a', header=False)
print(df.loc[5:20,["Unnamed: 5"]])
df.loc[1:1,["Unnamed: 0"]].to_csv(path_to_save + "test.csv", mode='a', header=False)
"""

# main function
def dataApp():

    if __name__ == '__main__':
        takeFolderList(path_folders)
        sort()
        listOfAllAdress()
        onlyExcel()
        exportCsv()

dataApp()
